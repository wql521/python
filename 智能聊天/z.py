def z(j):
    import openpyxl
    wb = openpyxl.load_workbook('智能聊天.xlsx')  # 打开xls文件
    ws = wb['sheet1']  # 获取表格
    # 获取表格的行数
    row = ws.max_row
    # 循环遍历表格
    Z=[]
    for i in range(1, row + 1):
        if i == j:
            # 获取表格的列数
            col = ws.max_column
            # 循环遍历表格
            for j in range(1, col + 1):
                # 获取表格的值
                Z.append(ws.cell(i, j).value)
    # 删除列表中的空值
    Z = list(filter(None, Z))
    # 删除列表第一个值
    Z.pop(0)
    return Z

