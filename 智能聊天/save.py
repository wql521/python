def save(me, f, r, b):
    import openpyxl
    wb = openpyxl.load_workbook('智能聊天.xlsx')  # 打开xls文件
    ws = wb['sheet1']  # 获取表格
    if r == 2 and b == -1:
        row = ws.max_row + 1
        ws.cell(row, 1).value = me
        ws.cell(row, 2).value = f
        wb.save('智能聊天.xlsx')
    else:
        ws.cell(b, r).value = f  # r为列 b为行
        wb.save('智能聊天.xlsx')
