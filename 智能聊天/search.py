def search(chat):
    import openpyxl
    import random
    import z
    wb = openpyxl.load_workbook('智能聊天.xlsx')  # 打开xls文件
    ws = wb['sheet1']  # 获取表格
    for i in range(1, ws.max_row + 1):  # 行
        if chat == ws.cell(i, 1).value:  # 查找是否有相同的内容
            q = z.z(i)
            if len(q) == 1:
                return q[0], i, 2  # 返回第一个值和行数
            else:
                a = random.randint(0, len(q) - 1)
                l = len(q)
                l = l + 1
                return q[a], i, l
        else:
            continue
    return '没有找到相关内容'
